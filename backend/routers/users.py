# ==========================================================
# PricePilot AI - User Management Router (Admin Only)
# Enterprise REST APIs for User CRUD & Approval Operations
# ==========================================================

import io
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from database import get_db
    from models import User, ActivityLog, Notification, Prediction
    from schemas import UserResponse, UserCreate, UserUpdateAdmin, BulkStatusRequest, BulkDeleteRequest
    from routers.auth import require_admin
    from security import get_password_hash
except ImportError:
    from backend.database import get_db
    from backend.models import User, ActivityLog, Notification, Prediction
    from backend.schemas import UserResponse, UserCreate, UserUpdateAdmin, BulkStatusRequest, BulkDeleteRequest
    from backend.routers.auth import require_admin
    from backend.security import get_password_hash

router = APIRouter(prefix="/api/users", tags=["User Management"])

# Secondary router for /api/admin endpoints as requested in Task 1
admin_router = APIRouter(prefix="/api/admin", tags=["Admin Management"])



# ==========================================================
# List All Users Endpoint (Admin Only)
# ==========================================================

@router.get("", response_model=List[UserResponse])
def get_all_users(
    status_filter: Optional[str] = None,
    db: Session = Depends(get_db),
    admin_user: User = Depends(require_admin)
):
    query = db.query(User)
    if status_filter:
        query = query.filter(User.status == status_filter.lower())
    return query.order_by(User.id.asc()).all()


# ==========================================================
# Create New User Endpoint (Admin Only)
# ==========================================================

@router.post("", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
def create_user(
    data: UserCreate,
    db: Session = Depends(get_db),
    admin_user: User = Depends(require_admin)
):
    clean_email = data.email.strip().lower()
    clean_username = data.username.strip()
    clean_name = data.name.strip()

    if db.query(User).filter(User.email == clean_email).first():
        raise HTTPException(status_code=400, detail="User with this email already exists.")

    if db.query(User).filter(User.username == clean_username).first():
        raise HTTPException(status_code=400, detail="Username is already taken.")

    role_formatted = "Admin" if data.role.strip().lower() in ["admin", "administrator"] else "User"

    new_user = User(
        name=clean_name,
        email=clean_email,
        username=clean_username,
        password_hash=get_password_hash(data.password),
        role=role_formatted,
        phone_number=data.phone_number,
        avatar_url=data.avatar_url,
        is_active=data.is_active,
        is_approved=data.is_approved,
        status=data.status
    )

    db.add(new_user)
    db.commit()
    db.refresh(new_user)

    log = ActivityLog(
        user_id=admin_user.id,
        action=f"Admin {admin_user.username} created user: {new_user.username} ({new_user.role})"
    )
    db.add(log)
    db.commit()

    return new_user


# ==========================================================
# User Approval & Status Endpoints (Admin Only)
# ==========================================================

@router.put("/{user_id}/approve", response_model=UserResponse)
def approve_user(
    user_id: int,
    db: Session = Depends(get_db),
    admin_user: User = Depends(require_admin)
):
    target_user = db.query(User).filter(User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found.")

    target_user.is_approved = True
    target_user.status = "approved"
    target_user.is_active = True
    db.commit()
    db.refresh(target_user)

    log = ActivityLog(
        user_id=admin_user.id,
        action=f"Admin {admin_user.username} approved user account: {target_user.username}"
    )
    db.add(log)
    db.commit()

    return target_user


@router.put("/{user_id}/reject", response_model=UserResponse)
def reject_user(
    user_id: int,
    db: Session = Depends(get_db),
    admin_user: User = Depends(require_admin)
):
    target_user = db.query(User).filter(User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found.")

    if target_user.id == admin_user.id:
        raise HTTPException(status_code=400, detail="Cannot reject your own admin account.")

    target_user.is_approved = False
    target_user.status = "rejected"
    target_user.is_active = False
    db.commit()
    db.refresh(target_user)

    log = ActivityLog(
        user_id=admin_user.id,
        action=f"Admin {admin_user.username} rejected user registration: {target_user.username}"
    )
    db.add(log)
    db.commit()

    return target_user


@router.put("/{user_id}/suspend", response_model=UserResponse)
def suspend_user(
    user_id: int,
    db: Session = Depends(get_db),
    admin_user: User = Depends(require_admin)
):
    target_user = db.query(User).filter(User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found.")

    if target_user.id == admin_user.id:
        raise HTTPException(status_code=400, detail="Cannot suspend your own admin account.")

    target_user.status = "suspended"
    target_user.is_active = False
    db.commit()
    db.refresh(target_user)

    log = ActivityLog(
        user_id=admin_user.id,
        action=f"Admin {admin_user.username} suspended user account: {target_user.username}"
    )
    db.add(log)
    db.commit()

    return target_user


@router.put("/{user_id}/role", response_model=UserResponse)
def change_user_role(
    user_id: int,
    new_role: str,
    db: Session = Depends(get_db),
    admin_user: User = Depends(require_admin)
):
    target_user = db.query(User).filter(User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found.")

    formatted_role = "Admin" if new_role.strip().lower() in ["admin", "administrator"] else "User"
    target_user.role = formatted_role
    db.commit()
    db.refresh(target_user)

    log = ActivityLog(
        user_id=admin_user.id,
        action=f"Admin {admin_user.username} changed role of {target_user.username} to {formatted_role}"
    )
    db.add(log)
    db.commit()

    return target_user


@router.put("/{user_id}/reset-password", response_model=UserResponse)
def admin_reset_password(
    user_id: int,
    new_password: str,
    db: Session = Depends(get_db),
    admin_user: User = Depends(require_admin)
):
    target_user = db.query(User).filter(User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found.")

    if len(new_password.strip()) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters.")

    target_user.password_hash = get_password_hash(new_password.strip())
    db.commit()
    db.refresh(target_user)

    log = ActivityLog(
        user_id=admin_user.id,
        action=f"Admin {admin_user.username} reset password for user: {target_user.username}"
    )
    db.add(log)
    db.commit()

    return target_user


# ==========================================================
# Update User Endpoint (Admin Only)
# ==========================================================

@router.put("/{user_id}", response_model=UserResponse)
def update_user(
    user_id: int,
    data: UserUpdateAdmin,
    db: Session = Depends(get_db),
    admin_user: User = Depends(require_admin)
):
    target_user = db.query(User).filter(User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found.")

    if data.name and data.name.strip():
        target_user.name = data.name.strip()

    if data.email and data.email.strip():
        clean_email = data.email.strip().lower()
        if clean_email != target_user.email:
            if db.query(User).filter(User.email == clean_email).first():
                raise HTTPException(status_code=400, detail="Email is already used by another account.")
            target_user.email = clean_email

    if data.username and data.username.strip():
        clean_username = data.username.strip()
        if clean_username != target_user.username:
            if db.query(User).filter(User.username == clean_username).first():
                raise HTTPException(status_code=400, detail="Username is already taken.")
            target_user.username = clean_username

    if data.role and data.role.strip():
        target_user.role = "Admin" if data.role.strip().lower() in ["admin", "administrator"] else "User"

    if data.phone_number is not None:
        target_user.phone_number = data.phone_number.strip() if data.phone_number.strip() else None

    if data.avatar_url is not None:
        target_user.avatar_url = data.avatar_url.strip() if data.avatar_url.strip() else None

    if data.is_active is not None:
        target_user.is_active = data.is_active

    if data.is_approved is not None:
        target_user.is_approved = data.is_approved
        if data.is_approved:
            target_user.status = "approved"

    if data.status and data.status.strip():
        target_user.status = data.status.strip()

    if data.password and data.password.strip():
        if len(data.password.strip()) < 6:
            raise HTTPException(status_code=400, detail="Password must be at least 6 characters.")
        target_user.password_hash = get_password_hash(data.password.strip())

    db.commit()
    db.refresh(target_user)

    log = ActivityLog(
        user_id=admin_user.id,
        action=f"Admin {admin_user.username} updated user #{target_user.id} ({target_user.username})"
    )
    db.add(log)
    db.commit()

    return target_user


# ==========================================================
# Delete User Endpoint (Admin Only)
# ==========================================================

@router.delete("/{user_id}", status_code=status.HTTP_200_OK)
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    admin_user: User = Depends(require_admin)
):
    if target_user := db.query(User).filter(User.id == user_id).first():
        # Prevent self deletion of active admin account
        if target_user.id == admin_user.id:
            raise HTTPException(status_code=400, detail="Cannot delete your own logged-in admin account.")

        username_deleted = target_user.username

        # Clean foreign key references in activity_logs
        db.query(ActivityLog).filter(ActivityLog.user_id == target_user.id).delete()
        db.delete(target_user)
        db.commit()

        log = ActivityLog(
            user_id=admin_user.id,
            action=f"Admin {admin_user.username} deleted user account: {username_deleted} (ID #{user_id})"
        )
        db.add(log)
        db.commit()

        return {"message": f"User {username_deleted} deleted successfully."}
    else:
        raise HTTPException(status_code=404, detail="User not found.")


# ==========================================================
# Excel User Export Generation Core Function
# ==========================================================

def generate_users_excel(
    user_ids: Optional[str],
    db: Session,
    admin_user: User
) -> StreamingResponse:
    query = db.query(User)
    if user_ids:
        ids_list = [int(i.strip()) for i in user_ids.split(",") if i.strip().isdigit()]
        if ids_list:
            query = query.filter(User.id.in_(ids_list))

    users_data = query.order_by(User.id.asc()).all()

    # Pre-calculate prediction counts per user
    pred_counts = dict(
        db.query(Prediction.user_id, func.count(Prediction.id))
        .filter(Prediction.user_id.isnot(None))
        .group_by(Prediction.user_id)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users Report"
    ws.views.sheetView[0].showGridLines = True

    # Styling definitions
    title_font = Font(name="Calibri", size=16, bold=True, color="1E3A8A")
    meta_font = Font(name="Calibri", size=10, italic=True, color="4B5563")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    even_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    # Row 1: Title
    ws.cell(row=1, column=1, value="PricePilot AI - Enterprise User Management Report").font = title_font

    # Row 2: Company Metadata & Timestamp
    gen_time = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    ws.cell(
        row=2,
        column=1,
        value=f"Company: PricePilot AI Enterprise | Exported By: {admin_user.username} ({admin_user.email}) | Timestamp: {gen_time} | Records: {len(users_data)}"
    ).font = meta_font

    # Headers
    headers = [
        "User ID", "Full Name", "Username", "Email", "Role",
        "Status", "Phone Number", "Created Date", "Updated Date",
        "Last Login", "Prediction Count", "OTP Verified", "Account Active"
    ]

    start_row = 4
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    ws.row_dimensions[start_row].height = 28

    # Populate Data Rows
    for row_idx, user in enumerate(users_data, start=start_row + 1):
        created_str = user.created_at.strftime("%Y-%m-%d %H:%M:%S") if user.created_at else "N/A"
        updated_str = user.updated_at.strftime("%Y-%m-%d %H:%M:%S") if user.updated_at else "N/A"
        last_login_str = user.last_login.strftime("%Y-%m-%d %H:%M:%S") if user.last_login else "Never"
        user_pred_count = pred_counts.get(user.id, 0)
        otp_verified_str = "Yes" if user.is_approved else "No"
        active_str = "Yes" if user.is_active else "No"

        row_data = [
            user.id,
            user.name or "",
            user.username or "",
            user.email or "",
            user.role or "User",
            (user.status or "active").capitalize(),
            user.phone_number or "N/A",
            created_str,
            updated_str,
            last_login_str,
            user_pred_count,
            otp_verified_str,
            active_str
        ]

        row_fill = even_row_fill if (row_idx % 2 == 0) else odd_row_fill

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = thin_border

            if col_idx in [1, 5, 6, 11, 12, 13]:
                cell.alignment = align_center
            elif col_idx in [8, 9, 10]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

        ws.row_dimensions[row_idx].height = 22

    # Enable Filter & Freeze Pane
    max_col_letter = get_column_letter(len(headers))
    end_row = start_row + max(len(users_data), 1)
    ws.auto_filter.ref = f"A{start_row}:{max_col_letter}{end_row}"
    ws.freeze_panes = f"A{start_row + 1}"

    # Auto Column Width Adjustment
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < start_row:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    headers_response = {
        'Content-Disposition': 'attachment; filename=Users_Report.xlsx'
    }
    return StreamingResponse(
        output,
        headers=headers_response,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ==========================================================
# Excel User Export Endpoints (Admin Only)
# ==========================================================

@router.get("/export/excel")
def export_users_excel_route(
    user_ids: Optional[str] = None,
    db: Session = Depends(get_db),
    admin_user: User = Depends(require_admin)
):
    return generate_users_excel(user_ids, db, admin_user)


@admin_router.get("/export-users")
def admin_export_users_route(
    user_ids: Optional[str] = None,
    db: Session = Depends(get_db),
    admin_user: User = Depends(require_admin)
):
    return generate_users_excel(user_ids, db, admin_user)


# ==========================================================
# Bulk User Operations Endpoints (Admin Only)
# ==========================================================

@router.post("/bulk-status")
def bulk_update_status(
    payload: BulkStatusRequest,
    db: Session = Depends(get_db),
    admin_user: User = Depends(require_admin)
):
    if not payload.user_ids:
        raise HTTPException(status_code=400, detail="No user IDs provided.")

    target_status = payload.status.strip().lower()
    is_act = target_status in ["approved", "active"]
    is_app = target_status == "approved" or (target_status == "active")

    # Update users
    users_to_update = db.query(User).filter(User.id.in_(payload.user_ids)).all()
    count = 0
    for u in users_to_update:
        if u.id == admin_user.id and target_status in ["suspended", "rejected", "inactive"]:
            continue  # Skip self suspension
        u.status = target_status
        u.is_active = is_act
        u.is_approved = is_app
        count += 1

    db.commit()

    log = ActivityLog(
        user_id=admin_user.id,
        action=f"Admin {admin_user.username} bulk updated status of {count} users to '{target_status}'"
    )
    db.add(log)
    db.commit()

    return {"message": f"Successfully updated {count} users to '{target_status}' status."}


@router.post("/bulk-delete")
def bulk_delete_users(
    payload: BulkDeleteRequest,
    db: Session = Depends(get_db),
    admin_user: User = Depends(require_admin)
):
    if not payload.user_ids:
        raise HTTPException(status_code=400, detail="No user IDs provided.")

    # Filter out admin self ID
    safe_ids = [uid for uid in payload.user_ids if uid != admin_user.id]
    if not safe_ids:
        raise HTTPException(status_code=400, detail="Cannot delete your own admin account.")

    # Delete related activity logs first
    db.query(ActivityLog).filter(ActivityLog.user_id.in_(safe_ids)).delete(synchronize_session=False)

    deleted_count = db.query(User).filter(User.id.in_(safe_ids)).delete(synchronize_session=False)
    db.commit()

    log = ActivityLog(
        user_id=admin_user.id,
        action=f"Admin {admin_user.username} bulk deleted {deleted_count} user accounts."
    )
    db.add(log)
    db.commit()

    return {"message": f"Successfully deleted {deleted_count} users."}


